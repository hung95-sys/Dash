"""
Module quản lý kết nối và thao tác với Google Sheets
"""
import gspread
import os
import tempfile
import threading
import time
from google.oauth2.service_account import Credentials
from config import (
    GOOGLE_SHEETS_CREDENTIALS_FILE,
    GOOGLE_SHEETS_SPREADSHEET_ID,
    SHEET_NAME_INCOME_EXPENSE,
    SHEET_NAME_CATEGORIES,
    SHEET_NAME_ACCOUNTS,
    CATEGORY_COLUMNS,
)
import unicodedata
from typing import Optional, Dict
from datetime import datetime
from gspread.utils import a1_to_rowcol

# Excel export helper (server-local copy)
try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    Workbook = None
    load_workbook = None

# Scope cho Google Sheets API
SCOPES = [
    'https://www.googleapis.com/auth/spreadsheets',
    'https://www.googleapis.com/auth/drive'
]


DEFAULT_HEADERS = ['Ngày', 'Loại', 'Danh mục', 'Số tiền', 'Ghi chú', 'Ngày tạo']


class SheetsManager:
    """Quản lý kết nối và thao tác với Google Sheets"""
    
    def __init__(self):
        # Chạy ở chế độ offline, chỉ đọc/ghi file xlsx trong thư mục data
        self.client = None
        self.spreadsheet = None
        self.local_only = True
        self.data_dir = os.path.join(os.path.dirname(__file__), 'data')
        if not os.path.exists(self.data_dir):
            os.makedirs(self.data_dir, exist_ok=True)
        self.local_full_path = os.path.join(self.data_dir, 'export_all.xlsx')
        self._local_wb = None
        self._local_wb_mtime = 0
        self._records_cache = {}  # Cache records theo sheet_name để tránh parse lại
        # Offline mode: bỏ qua kết nối Google Sheets
        if not self.local_only:
            self._connect()

    # ---------- Logging helper ----------
    def _log_source(self, sheet_name: str, source: str):
        try:
            print(f"[DATA] {sheet_name}: source={source}")
        except Exception:
            pass
    
    def _connect(self):
        """Kết nối với Google Sheets"""
        try:
            creds = Credentials.from_service_account_file(
                GOOGLE_SHEETS_CREDENTIALS_FILE,
                scopes=SCOPES
            )
            self.client = gspread.authorize(creds)
            if GOOGLE_SHEETS_SPREADSHEET_ID:
                self.spreadsheet = self.client.open_by_key(GOOGLE_SHEETS_SPREADSHEET_ID)
            else:
                # Tạo spreadsheet mới nếu chưa có ID
                self.spreadsheet = self.client.create('Personal Finance Manager')
                print(f"Đã tạo spreadsheet mới: {self.spreadsheet.id}")
        except Exception as e:
            print(f"Lỗi kết nối Google Sheets: {e}")
            raise

    # ---------- Local helpers ----------
    def _ensure_local_file(self):
        """Đảm bảo file xlsx local tồn tại."""
        if Workbook is None:
            return False
        if not os.path.exists(self.local_full_path):
            wb = Workbook()
            # Xóa sheet mặc định để giữ tên sheet theo dữ liệu import
            wb.remove(wb.active)
            wb.save(self.local_full_path)
        return True

    def _get_local_values(self, sheet_name: str):
        """Lấy tất cả giá trị dạng list[list] từ file local."""
        ws = self._get_local_ws(sheet_name, read_only=True)
        if not ws:
            return []
        return [list(row) for row in ws.iter_rows(values_only=True)]
    
    def get_sheet(self, sheet_name):
        """Lấy worksheet theo tên"""
        try:
            return self.spreadsheet.worksheet(sheet_name)
        except gspread.exceptions.WorksheetNotFound:
            # Tạo worksheet mới nếu chưa tồn tại
            return self.spreadsheet.add_worksheet(title=sheet_name, rows=1000, cols=20)

    # ---------- Local workbook helpers ----------
    def _load_local_workbook(self, read_only=True):
        """Nạp workbook local nếu tồn tại, cache theo mtime."""
        if Workbook is None:
            return None
        # Tự tạo file rỗng nếu chưa có để tránh fallback Google Sheets
        self._ensure_local_file()
        mtime = os.path.getmtime(self.local_full_path)
        if self._local_wb and mtime == self._local_wb_mtime and (not read_only or hasattr(self._local_wb, '_read_only')):
            return self._local_wb
        try:
            # Dùng read_only=True để đọc nhanh hơn, không load công thức
            wb = load_workbook(self.local_full_path, read_only=read_only, data_only=True)
            if read_only:
                self._local_wb = wb
                self._local_wb_mtime = mtime
            # Clear cache records khi file thay đổi
            self._records_cache.clear()
            return wb
        except Exception as e:
            print(f"[LOCAL] Lỗi đọc local workbook: {e}")
            if read_only:
                self._local_wb = None
                self._records_cache.clear()
            return None

    def _load_local_workbook_for_write(self):
        """Load workbook local ở chế độ có thể ghi."""
        if Workbook is None:
            return None
        if not os.path.exists(self.local_full_path):
            # Tạo file mới nếu chưa có
            wb = Workbook()
            wb.remove(wb.active)  # Xóa sheet mặc định
            wb.save(self.local_full_path)
        return load_workbook(self.local_full_path, read_only=False, data_only=True)

    def _get_local_ws(self, sheet_name: str, read_only: bool = True, create: bool = False, headers: list = None):
        """Lấy worksheet từ file local, so khớp không dấu/không khoảng; tạo mới nếu cần."""
        wb = self._load_local_workbook(read_only=read_only) if read_only else self._load_local_workbook_for_write()
        if not wb:
            return None
        target_norm = self._normalize_title(sheet_name)
        ws = None
        for name in wb.sheetnames:
            if self._normalize_title(name) == target_norm:
                ws = wb[name]
                break
        if not ws and create:
            ws = self._get_or_create_sheet_in_workbook(wb, sheet_name, headers)
            if not read_only:
                self._save_local_workbook(wb)
        return ws

    def _save_local_workbook(self, wb):
        """Lưu workbook vào file local và invalidate cache."""
        if wb is None:
            return
        try:
            wb.save(self.local_full_path)
            # Invalidate cache
            self._local_wb = None
            self._local_wb_mtime = 0
            self._records_cache.clear()
            print(f"[LOCAL] Đã lưu workbook vào {self.local_full_path}")
        except Exception as e:
            print(f"[LOCAL] Lỗi lưu workbook: {e}")
            raise

    def _get_records_from_local(self, sheet_name: str):
        """Đọc records từ file local nếu có, cache trong memory."""
        import time
        start = time.time()
        # Kiểm tra cache trước
        cache_key = f"{sheet_name}_{self._local_wb_mtime}"
        if cache_key in self._records_cache:
            elapsed = (time.time() - start) * 1000
            print(f"[PERF] {sheet_name}: cache hit ({elapsed:.1f}ms)")
            return self._records_cache[cache_key]
        
        ws = self._get_local_ws(sheet_name, read_only=True)
        if not ws:
            return []
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            self._records_cache[cache_key] = []
            return []
        headers = [str(h).strip() if h else '' for h in rows[0]]
        records = []
        for row in rows[1:]:
            rec = {}
            for idx, h in enumerate(headers):
                if not h:
                    continue
                rec[h] = row[idx] if idx < len(row) else ""
            records.append(rec)
        # Cache lại
        self._records_cache[cache_key] = records
        elapsed = (time.time() - start) * 1000
        print(f"[PERF] {sheet_name}: parsed {len(records)} rows ({elapsed:.1f}ms)")
        return records

    def _get_or_create_sheet_in_workbook(self, wb, sheet_name: str, headers: list = None):
        """Lấy hoặc tạo sheet trong workbook."""
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
            if headers:
                ws.append(headers)
        return ws

    def _add_row_to_local_sheet(self, sheet_name: str, row_data: list, headers: list = None):
        """Thêm một dòng vào sheet trong file local."""
        wb = self._load_local_workbook_for_write()
        if not wb:
            raise RuntimeError("Không thể mở file local để ghi")
        ws = self._get_or_create_sheet_in_workbook(wb, sheet_name, headers)
        ws.append(row_data)
        self._save_local_workbook(wb)

    def _update_row_in_local_sheet(self, sheet_name: str, row_index: int, row_data: list):
        """Cập nhật một dòng trong sheet của file local (row_index bắt đầu từ 1, 1 là header)."""
        wb = self._load_local_workbook_for_write()
        if not wb:
            raise RuntimeError("Không thể mở file local để ghi")
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' không tồn tại")
        ws = wb[sheet_name]
        if row_index < 1 or row_index > ws.max_row:
            raise ValueError(f"Row index {row_index} không hợp lệ")
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_index, column=col_idx, value=value)
        self._save_local_workbook(wb)

    def _delete_row_in_local_sheet(self, sheet_name: str, row_index: int):
        """Xóa một dòng trong sheet của file local (row_index bắt đầu từ 1, 1 là header)."""
        wb = self._load_local_workbook_for_write()
        if not wb:
            raise RuntimeError("Không thể mở file local để ghi")
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' không tồn tại")
        ws = wb[sheet_name]
        if row_index < 1 or row_index > ws.max_row:
            raise ValueError(f"Row index {row_index} không hợp lệ")
        ws.delete_rows(row_index)
        self._save_local_workbook(wb)

    def _get_records_local_first(self, sheet_name: str):
        """Ưu tiên đọc local, nếu không có thì đọc Google Sheet."""
        local_records = self._get_records_from_local(sheet_name)
        # Ở chế độ offline luôn dùng file local, nếu không có trả về rỗng
        return local_records or []

    # ---------- Export toàn bộ spreadsheet thành 1 file xlsx ----------
    def export_full_spreadsheet(self, dest_path: str):
        """
        Sao chép toàn bộ các worksheet trong Google Sheet hiện tại vào một file xlsx duy nhất.
        """
        if Workbook is None:
            raise RuntimeError("Thiếu thư viện openpyxl để xuất file")
        # Offline mode: sao chép file local hiện có
        if not os.path.exists(self.local_full_path):
            raise FileNotFoundError(f"Không tìm thấy file local: {self.local_full_path}")
        import shutil
        shutil.copy(self.local_full_path, dest_path)
        return dest_path

    def sync_local_to_sheet(self, local_path: Optional[str] = None):
        """Ghi đè toàn bộ Google Sheet bằng dữ liệu từ file local (toàn bộ workbook)."""
        raise RuntimeError("Chế độ offline: sync_local_to_sheet không khả dụng")
    
    def import_full_spreadsheet(self, source_path: str):
        """Import file xlsx từ đường dẫn source_path và ghi đè lên file local hiện tại."""
        if Workbook is None:
            raise RuntimeError("Thiếu thư viện openpyxl để import file")
        if not os.path.exists(source_path):
            raise FileNotFoundError(f"Không tìm thấy file: {source_path}")
        
        try:
            # Sao chép file import vào vị trí file local
            import shutil
            shutil.copy(source_path, self.local_full_path)
            
            # Invalidate cache
            self._local_wb = None
            self._local_wb_mtime = 0
            self._records_cache.clear()
            
            print(f"[LOCAL] Đã import file từ {source_path} vào {self.local_full_path}")
            return True
        except Exception as e:
            print(f"[LOCAL] Lỗi import file: {e}")
            raise
    def init_income_expense_sheet(self):
        """Khởi tạo sheet ThuChi với headers (linh hoạt theo sheet hiện có)"""
        wb = self._load_local_workbook_for_write()
        if not wb:
            raise RuntimeError("Không thể mở file local để khởi tạo ThuChi")
        ws = self._get_or_create_sheet_in_workbook(wb, SHEET_NAME_INCOME_EXPENSE, DEFAULT_HEADERS)
        self._save_local_workbook(wb)
        return ws

    # ---------- Accounts (Users) ----------
    def _ensure_accounts_headers(self, sheet):
        headers = sheet.row_values(1)
        # Chuẩn mới dùng 'user' thay vì 'username'
        expected = ['user', 'password', 'Name', 'role', 'active']
        if not headers:
            sheet.append_row(expected)
            return
        # Nếu còn header cũ, chuyển sang header mới
        if headers and headers[0].strip().lower() == 'username':
            headers[0] = 'user'
        if len(headers) < len(expected):
            headers = headers + expected[len(headers):]
        sheet.update("A1:E1", [headers])

    def _normalize_title(self, title: str) -> str:
        """Chuẩn hóa tên sheet để so khớp: lower, bỏ dấu, bỏ khoảng trắng."""
        if not title:
            return ''
        nfkd = unicodedata.normalize('NFD', title)
        no_accent = ''.join(c for c in nfkd if unicodedata.category(c) != 'Mn')
        return no_accent.replace(' ', '').lower()

    def get_categories_sheet(self):
        """Lấy hoặc tạo sheet Danh Mục trong file local (offline)."""
        headers = ['Danh mục Thu', 'Danh mục Chi', 'Mục đích Quỹ', 'Icon D', 'Icon E']
        ws = self._get_local_ws(SHEET_NAME_CATEGORIES, read_only=False, create=True, headers=headers)
        if not ws:
            raise RuntimeError("Không thể mở file local để lấy sheet Danh Mục")
        return ws

    def init_accounts_sheet(self):
        sheet = self.get_sheet(SHEET_NAME_ACCOUNTS)
        self._ensure_accounts_headers(sheet)
        return sheet

    def get_accounts(self):
        """Trả về danh sách tài khoản với row_number để CRUD."""
        local_rows = self._get_records_from_local(SHEET_NAME_ACCOUNTS)
        sheet = None
        if local_rows is not None:
            headers = ['user', 'password', 'Name', 'role', 'active']
            rows = [headers] + [[r.get(h, "") for h in headers] for r in local_rows]
            self._log_source(SHEET_NAME_ACCOUNTS, 'local')
        else:
            sheet = self.init_accounts_sheet()
            rows = sheet.get_all_values()
            self._log_source(SHEET_NAME_ACCOUNTS, 'sheet')
        accounts = []
        for idx, row in enumerate(rows[1:], start=2):
            if not row or all(not str(c).strip() for c in row):
                continue
            user_val = (row[0] if len(row) > 0 else '').strip()
            password = (row[1] if len(row) > 1 else '').strip()
            full_name = (row[2] if len(row) > 2 else '').strip()
            role = (row[3] if len(row) > 3 else 'user').strip() or 'user'
            active_raw = (row[4] if len(row) > 4 else '').strip().upper()
            active = active_raw == 'TRUE'
            if not user_val:
                continue
            accounts.append({
                'row_number': idx,
                'username': user_val,  # giữ key 'username' cho compat API/JS
                'user': user_val,
                'password': password,
                'name': full_name,
                'role': role,
                'active': active,
            })
        return accounts

    def get_account_by_username(self, username):
        username = (username or '').strip()
        if not username:
            return None
        for acc in self.get_accounts():
            if acc['username'] == username:
                return acc
        return None

    def add_account(self, username, password, name='', role='user', active=True):
        """Thêm tài khoản vào file local."""
        username = (username or '').strip()
        password = (password or '').strip()
        name = (name or '').strip()
        role = (role or 'user').strip() or 'user'
        if not username or not password:
            raise ValueError('Username và password không được trống')
        if role not in ['admin', 'user']:
            raise ValueError('Role phải là admin hoặc user')

        # kiểm tra trùng
        if self.get_account_by_username(username):
            raise ValueError('Tài khoản đã tồn tại')

        # Tính next_row từ file local
        accounts = self.get_accounts()
        next_row = len(accounts) + 2  # +2 vì có header và index bắt đầu từ 2

        headers = ['user', 'password', 'Name', 'role', 'active']
        row = [username, password, name, role, 'TRUE' if active else 'FALSE']
        self._add_row_to_local_sheet(SHEET_NAME_ACCOUNTS, row, headers)

        # Tự động tạo worksheet trống theo Name (nếu có) trong file local
        if name:
            try:
                wb = self._load_local_workbook_for_write()
                if wb:
                    ws = self._get_or_create_sheet_in_workbook(wb, name, ['Ngày', 'Loại', 'Danh mục', 'Số tiền', 'Ghi chú', 'Quỹ'])
                    self._save_local_workbook(wb)
            except Exception as e:
                print(f"[DEBUG] Không thể tạo sheet cho Name='{name}': {e}")
        
        print(f"[LOCAL] Đã thêm tài khoản '{username}' vào sheet '{SHEET_NAME_ACCOUNTS}'")
        return next_row

    def update_password_by_username(self, username, new_password):
        username = (username or '').strip()
        new_password = (new_password or '').strip()
        if not username or not new_password:
            raise ValueError('Thiếu username hoặc mật khẩu')
        accounts = self.get_accounts()
        for acc in accounts:
            if acc['username'] == username:
                row = acc['row_number']
                return self.update_account(row, password=new_password)
        raise ValueError('Không tìm thấy tài khoản')

    # ---------- User sheet helpers ----------
    def get_user_sheet_name(self, username: str) -> Optional[str]:
        acc = self.get_account_by_username(username)
        if acc:
            return acc.get('name') or acc.get('username')
        return None

    def get_user_summary(self, username: str) -> Dict[str, float]:
        sheet_name = self.get_user_sheet_name(username)
        if not sheet_name:
            return {'total_income': 0.0, 'total_expense': 0.0, 'balance': 0.0, 'total_fund': 0.0}
        
        # Ưu tiên đọc từ local
        records = self._get_records_from_local(sheet_name)
        if records is not None:
            self._log_source(sheet_name, 'local')
            rows = []
            # Convert records thành rows format
            if records:
                headers = list(records[0].keys()) if records else ['Ngày', 'Loại', 'Danh mục', 'Số tiền', 'Ghi chú', 'Quỹ']
                rows.append(headers)
                for r in records:
                    rows.append([r.get(h, '') for h in headers])
        else:
            sheet = self.get_sheet(sheet_name)
            self._ensure_user_sheet_headers(sheet)
            rows = sheet.get_all_values()
            self._log_source(sheet_name, 'sheet')
        
        total_income = 0.0
        total_expense = 0.0
        total_fund = 0.0

        now = datetime.now()
        cur_month = now.month
        cur_year = now.year

        for row in rows[1:]:
            if not row or len(row) < 4:
                continue
            tx_type = (row[1] or '').strip()
            category = (row[2] or '').strip()
            amount_raw = str(row[3] or '').replace('.', '').replace(',', '').strip()
            try:
                amount = float(amount_raw)
            except (ValueError, TypeError):
                continue
            if amount <= 0:
                continue

            # Parse date to filter theo tháng hiện tại
            date_str = (row[0] or '').strip()
            tx_month = cur_month
            tx_year = cur_year
            if date_str:
                try:
                    tx_date = datetime.fromisoformat(date_str)
                    tx_month = tx_date.month
                    tx_year = tx_date.year
                except ValueError:
                    # Thử dd/mm/yyyy
                    try:
                        tx_date = datetime.strptime(date_str, '%d/%m/%Y')
                        tx_month = tx_date.month
                        tx_year = tx_date.year
                    except ValueError:
                        pass

            # Tổng Thu/Chi theo tháng hiện tại
            if tx_month == cur_month and tx_year == cur_year:
                if tx_type == 'Thu':
                    total_income += amount
                elif tx_type == 'Chi':
                    total_expense += amount

        # Tổng Quỹ lấy từ cả 2 sheet "Hưng" và "Giang" (không phải từ sheet của user hiện tại)
        quy_summary = self.get_combined_quy_summary(['Hưng', 'Giang'])
        total_fund = quy_summary.get('total', 0.0)

        balance = total_income - total_expense
        return {
            'total_income': total_income,
            'total_expense': total_expense,
            'balance': balance,
            'total_fund': total_fund,
        }

    def _ensure_user_sheet_headers(self, sheet):
        headers = sheet.row_values(1)
        expected = ['Ngày', 'Loại', 'Danh mục', 'Số tiền', 'Ghi chú', 'Quỹ']
        if not headers:
            sheet.append_row(expected)
            return
        if len(headers) < len(expected):
            headers = headers + expected[len(headers):]
        sheet.update("A1:F1", [headers[:6]])

    def add_user_transaction(self, sheet_name, date_str, category, amount, note='', is_income=False, purpose=''):
        """Thêm giao dịch vào file local. Nếu category='Quỹ' thì purpose sẽ được lưu vào cột 'Quỹ'."""
        if not sheet_name:
            raise ValueError('Thiếu tên sheet người dùng')
        category = (category or '').strip()
        if not category:
            raise ValueError('Danh mục không được trống')
        try:
            amount_val = float(amount)
            if amount_val <= 0:
                raise ValueError('Số tiền phải lớn hơn 0')
        except (ValueError, TypeError):
            raise ValueError('Số tiền không hợp lệ')

        note = (note or '').strip()
        purpose = (purpose or '').strip()
        tx_type = 'Thu' if is_income else 'Chi'
        headers = ['Ngày', 'Loại', 'Danh mục', 'Số tiền', 'Ghi chú', 'Quỹ']
        # Nếu category='Quỹ' thì lưu purpose vào cột 'Quỹ', ngược lại để trống
        quy_value = purpose if category.lower() == 'quỹ' else ''
        row = [date_str, tx_type, category, amount_val, note, quy_value]
        self._add_row_to_local_sheet(sheet_name, row, headers)
        print(f"[LOCAL] Đã thêm giao dịch vào sheet '{sheet_name}'")
        return True

    def get_user_transactions(self, sheet_name, month, year):
        """Lấy giao dịch theo tháng/năm, ưu tiên đọc file local (data/export_all.xlsx)."""
        if not sheet_name:
            raise ValueError('Thiếu tên sheet người dùng')
        try:
            month = int(month)
            year = int(year)
        except (ValueError, TypeError):
            raise ValueError('Tháng/năm không hợp lệ')

        records = self._get_records_from_local(sheet_name)
        if records is not None:
            self._log_source(sheet_name, 'local')
        else:
            sheet = self.get_sheet(sheet_name)
            self._ensure_user_sheet_headers(sheet)
            records = sheet.get_all_records()
            self._log_source(sheet_name, 'sheet')
            # lưu cache xlsx nếu có openpyxl
            try:
                self._save_cache_records(sheet_name, records)
            except Exception:
                pass
        txs = []
        for r in records:
            date_str = str(r.get('Ngày') or '').strip()
            if not date_str:
                continue
            try:
                dt = datetime.strptime(date_str, '%Y-%m-%d')
            except ValueError:
                # thử định dạng khác dd/mm/yyyy
                try:
                    dt = datetime.strptime(date_str, '%d/%m/%Y')
                except ValueError:
                    continue
            if dt.month != month or dt.year != year:
                continue
            txs.append({
                'date': dt.strftime('%Y-%m-%d'),
                'type': (r.get('Loại') or '').strip() or 'Chi',
                'category': (r.get('Danh mục') or '').strip(),
                'amount': float(r.get('Số tiền') or 0),
                'note': (r.get('Ghi chú') or '').strip(),
                'purpose': (r.get('Quỹ') or '').strip(),  # Mục đích quỹ từ cột "Quỹ"
            })
        return txs

    def get_user_yearly_report(self, sheet_name, years=5):
        """Tổng hợp thu/chi/quỹ theo năm, ưu tiên đọc local, nếu không có sẽ đọc Sheet rồi lưu cache."""
        if not sheet_name:
            raise ValueError('Thiếu tên sheet người dùng')
        try:
            years = int(years)
        except (ValueError, TypeError):
            years = 5
        years = max(1, min(10, years))  # giới hạn 1-10

        records = self._get_records_from_local(sheet_name)
        if records is not None:
            self._log_source(sheet_name, 'local')
        else:
            sheet = self.get_sheet(sheet_name)
            self._ensure_user_sheet_headers(sheet)
            records = sheet.get_all_records()
            self._log_source(sheet_name, 'sheet')
            try:
                self._save_cache_records(sheet_name, records)
            except Exception:
                pass

        agg = {}
        for r in records:
            date_str = str(r.get('Ngày') or '').strip()
            if not date_str:
                continue
            try:
                dt = datetime.strptime(date_str, '%Y-%m-%d')
            except ValueError:
                try:
                    dt = datetime.strptime(date_str, '%d/%m/%Y')
                except ValueError:
                    continue
            year = dt.year
            tx_type = (r.get('Loại') or '').strip()
            category = (r.get('Danh mục') or '').strip()
            amount = 0
            try:
                amount = float(r.get('Số tiền') or 0)
            except (ValueError, TypeError):
                continue
            if amount <= 0:
                continue
            if year not in agg:
                agg[year] = {'income': 0.0, 'expense': 0.0, 'fund': 0.0}
            if tx_type == 'Thu':
                agg[year]['income'] += amount
            elif tx_type == 'Chi':
                agg[year]['expense'] += amount
                if category.lower() == 'quỹ':
                    agg[year]['fund'] += amount

        # lấy các năm gần nhất, sắp xếp giảm dần rồi cắt
        years_sorted = sorted(agg.keys(), reverse=True)[:years]
        years_sorted.sort()  # hiển thị tăng dần cho chart
        return [{'year': y,
                 'income': agg[y].get('income', 0.0),
                 'expense': agg[y].get('expense', 0.0),
                 'fund': agg[y].get('fund', 0.0)} for y in years_sorted]

    def get_user_monthly_report(self, sheet_name, year=None):
        """Tổng hợp thu/chi/quỹ theo tháng trong một năm, ưu tiên đọc local."""
        if not sheet_name:
            raise ValueError('Thiếu tên sheet người dùng')
        try:
            year = int(year) if year else datetime.now().year
        except (ValueError, TypeError):
            year = datetime.now().year

        records = self._get_records_from_local(sheet_name)
        if records is None:
            sheet = self.get_sheet(sheet_name)
            self._ensure_user_sheet_headers(sheet)
            records = sheet.get_all_records()

        agg = {m: {'income': 0.0, 'expense': 0.0, 'fund': 0.0} for m in range(1, 13)}
        for r in records:
            date_str = str(r.get('Ngày') or '').strip()
            if not date_str:
                continue
            try:
                dt = datetime.strptime(date_str, '%Y-%m-%d')
            except ValueError:
                try:
                    dt = datetime.strptime(date_str, '%d/%m/%Y')
                except ValueError:
                    continue
            if dt.year != year:
                continue
            tx_type = (r.get('Loại') or '').strip()
            category = (r.get('Danh mục') or '').strip()
            try:
                amount = float(r.get('Số tiền') or 0)
            except (ValueError, TypeError):
                continue
            if amount <= 0:
                continue
            if tx_type == 'Thu':
                agg[dt.month]['income'] += amount
            elif tx_type == 'Chi':
                agg[dt.month]['expense'] += amount
                if category.lower() == 'quỹ':
                    agg[dt.month]['fund'] += amount

        return [{'month': m, **agg[m]} for m in range(1, 13)]

    def update_account(self, row_number, password=None, name=None, role=None, active=None):
        """Cập nhật tài khoản trong file local."""
        try:
            row_number = int(row_number)
        except (ValueError, TypeError):
            raise ValueError('ID tài khoản không hợp lệ')
        if row_number <= 1:
            raise ValueError('Không thể sửa header')

        # Đọc row hiện tại từ local
        accounts = self.get_accounts()
        acc = next((a for a in accounts if a['row_number'] == row_number), None)
        if not acc:
            raise ValueError('Không tìm thấy tài khoản')

        row = [acc['user'], acc['password'], acc['name'], acc['role'], 'TRUE' if acc['active'] else 'FALSE']
        if len(row) < 5:
            row += [''] * (5 - len(row))

        if password is not None and str(password).strip():
            row[1] = str(password).strip()
        if name is not None:
            row[2] = str(name).strip()
        if role is not None:
            role = role.strip() or 'user'
            if role not in ['admin', 'user']:
                raise ValueError('Role phải là admin hoặc user')
            row[3] = role
        if active is not None:
            row[4] = 'TRUE' if bool(active) else 'FALSE'

        self._update_row_in_local_sheet(SHEET_NAME_ACCOUNTS, row_number, row[:5])
        print(f"[LOCAL] Đã cập nhật tài khoản row {row_number}")
        return True

    def delete_account(self, row_number):
        """Xóa tài khoản từ file local."""
        try:
            row_number = int(row_number)
        except (ValueError, TypeError):
            raise ValueError('ID tài khoản không hợp lệ')
        if row_number <= 1:
            raise ValueError('Không thể xóa header')
        
        # Lấy Name (cột C) trước khi xóa để xóa worksheet tương ứng
        accounts = self.get_accounts()
        acc = next((a for a in accounts if a['row_number'] == row_number), None)
        if not acc:
            raise ValueError('Không tìm thấy tài khoản')
        name_value = (acc.get('name') or '').strip()

        # Xóa row trong file local
        self._delete_row_in_local_sheet(SHEET_NAME_ACCOUNTS, row_number)

        # Xóa worksheet trùng với Name nếu có và không phải sheet hệ thống
        protected_titles = {
            SHEET_NAME_ACCOUNTS.lower(),
            SHEET_NAME_CATEGORIES.lower(),
            SHEET_NAME_INCOME_EXPENSE.lower(),
        }
        if name_value:
            try:
                wb = self._load_local_workbook_for_write()
                if wb and name_value in wb.sheetnames:
                    if name_value.lower() not in protected_titles:
                        wb.remove(wb[name_value])
                        self._save_local_workbook(wb)
                        print(f"[LOCAL] Đã xóa sheet '{name_value}' do xóa tài khoản")
                    else:
                        print(f"[DEBUG] Bỏ qua không xóa sheet bảo vệ: {name_value}")
            except Exception as e:
                print(f"[DEBUG] Không thể xóa sheet '{name_value}': {e}")
        
        print(f"[LOCAL] Đã xóa tài khoản row {row_number}")
        return True

    def _ensure_headers(self, sheet):
        """Đảm bảo sheet có header hợp lệ"""
        existing_data = sheet.get_all_values()
        if not existing_data:
            sheet.append_row(DEFAULT_HEADERS)
            print(f"[DEBUG] Sheet '{sheet.title}' rỗng -> thêm header mặc định")
        else:
            header = existing_data[0]
            # Nếu header hiện tại có ít hơn 3 cột, thay bằng header mặc định
            if len(header) < 3 or header[0].strip() == '':
                sheet.insert_row(DEFAULT_HEADERS, 1)
                print(f"[DEBUG] Sheet '{sheet.title}' không có header -> chèn header ở dòng 1")

    def _get_headers(self, sheet):
        """Lấy danh sách headers hiện có"""
        headers = sheet.row_values(1)
        if not headers:
            headers = DEFAULT_HEADERS
            sheet.insert_row(headers, 1)
        return headers
    
    def add_transaction(self, transaction_type, category, amount, note='', purpose='', is_wife=False):
        """Thêm giao dịch thu/chi.

        Nếu Chi + Danh mục = 'Quỹ' và is_wife=True:
        - Ghi chú sẽ tự động là 'Vợ' (nếu bạn không nhập gì).
        - Giao dịch này được bỏ qua khi tính Tổng Chi.
        """
        # Validate dữ liệu
        transaction_type = (transaction_type or '').strip()
        if transaction_type not in ['Thu', 'Chi']:
            raise ValueError('Loại giao dịch không hợp lệ')

        category = (category or '').strip()
        if not category:
            raise ValueError('Danh mục không được để trống')

        try:
            amount = float(amount)
            if amount <= 0:
                raise ValueError('Số tiền phải lớn hơn 0')
            if amount > 1e12:
                raise ValueError('Số tiền quá lớn')
        except (ValueError, TypeError):
            raise ValueError('Số tiền không hợp lệ')

        purpose = (purpose or '').strip()
        is_wife_flag = bool(is_wife)

        # Nếu là chi cho vợ mà bạn chưa ghi chú gì, tự điền 'Vợ'
        if is_wife_flag and not note.strip():
            note = 'Vợ'

        # Với layout chuẩn:
        # A: Ngày, B: Loại, C: Danh mục, D: Số tiền, E: Ghi chú, F: Mục đích
        current_values = self._get_local_values(SHEET_NAME_INCOME_EXPENSE)
        next_row_number = len(current_values) + 1 if current_values else 1

        today = datetime.now().strftime('%Y-%m-%d')
        amount_formatted = f"{amount:.2f}" if amount < 1e10 else str(int(amount))

        headers = ['Ngày', 'Loại', 'Danh mục', 'Số tiền', 'Ghi chú', 'Mục đích']
        row = [today, transaction_type, category, amount_formatted, note.strip() if note else '', purpose]
        self._add_row_to_local_sheet(SHEET_NAME_INCOME_EXPENSE, row, headers)
        return next_row_number
    
    def _is_header_row(self, row):
        """Kiểm tra xem dòng có phải là header hay không (hỗ trợ có dấu / không dấu)."""
        if not row:
            return False

        # Ghép toàn bộ dòng lại để tìm từ khóa
        joined = " ".join(str(cell).upper().replace(" ", "") for cell in row if cell)
        if not joined:
            return False

        header_keywords = [
            "ID",
            "NGAY", "NGÀY",
            "LOAI", "LOẠI",
            "DANHMUC", "DANHMỤC",
            "SOTIEN", "SỐTIỀN",
            "GHICHU", "GHICHÚ",
        ]
        return any(kw in joined for kw in header_keywords)
    
    def _is_data_row(self, row):
        """Kiểm tra xem dòng có phải là dữ liệu không"""
        if not row or len(row) < 3:
            return False
        
        # Dòng dữ liệu thường có:
        # - Cột 0: ID (số)
        # - Cột 1: Ngày (format YYYY-MM-DD hoặc tương tự)
        # - Cột 2: Loại (Thu hoặc Chi)
        try:
            # Kiểm tra ID là số
            if row[0]:
                int(str(row[0]).strip())
            
            # Kiểm tra Loại là Thu hoặc Chi
            if len(row) > 2 and row[2]:
                type_str = str(row[2]).strip()
                if type_str in ['Thu', 'Chi']:
                    return True
        except (ValueError, TypeError):
            pass
        
        return False
    
    def get_all_transactions(self):
        """Lấy tất cả giao dịch"""
        rows = self._get_local_values(SHEET_NAME_INCOME_EXPENSE)
        print(f"[DEBUG] Đang đọc dữ liệu từ file local: '{SHEET_NAME_INCOME_EXPENSE}', tổng {len(rows)} dòng")

        if not rows:
            return []

        # Xác định header
        start_idx = 0
        if self._is_header_row(rows[0]):
            start_idx = 1

        headers = rows[0] if rows else DEFAULT_HEADERS
        header_map = [str(h).strip() for h in headers]

        transactions = []
        for idx, row in enumerate(rows[start_idx:], start=start_idx + 1):
            if not row or all((str(cell or '').strip() == '' for cell in row)):
                continue

            row_dict = {}
            for col_idx, header in enumerate(header_map):
                row_dict[header] = row[col_idx] if col_idx < len(row) else ''

            def get_value(possible_headers, fallback_index=None):
                for h in possible_headers:
                    if h in row_dict and row_dict[h]:
                        return str(row_dict[h]).strip()
                if fallback_index is not None and fallback_index < len(row):
                    return str(row[fallback_index] or '').strip()
                return ''

            date_value = get_value(['Ngày', 'Ngay', 'Date'], 0)
            type_value = get_value(['Loại', 'Loai', 'Type'], 1)
            category_value = get_value(['Danh mục', 'Danh muc', 'Category'], 2)
            amount_value = get_value(['Số tiền', 'So tien', 'Amount'], 3)
            note_value = get_value(['Ghi chú', 'Ghi chu', 'Note'], 4)
            purpose_value = get_value(['Mục đích', 'Muc dich', 'MucDich', 'Quỹ', 'Quy'], 5)

            if type_value not in ['Thu', 'Chi']:
                continue

            amount_clean = str(amount_value).replace(',', '').replace(' ', '')
            try:
                amount = float(amount_clean)
                if amount <= 0 or amount > 1e15:
                    continue
            except (ValueError, TypeError):
                continue

            transaction = {
                'id': str(idx),
                'row_number': idx,
                'date': date_value,
                'type': type_value,
                'category': category_value,
                'amount': amount,
                'note': note_value,
                'purpose': purpose_value,
                'is_wife': any(k in str(note_value).lower() for k in ['vợ', 'vo']),
            }
            transactions.append(transaction)

        return transactions
    
    def delete_transaction(self, row_number):
        """Xóa giao dịch theo số dòng trong sheet"""
        try:
            row_number = int(row_number)
        except (ValueError, TypeError):
            return False

        if row_number <= 1:
            # không xóa header
            return False

        try:
            self._delete_row_in_local_sheet(SHEET_NAME_INCOME_EXPENSE, row_number)
            print(f"[DEBUG] Đã xóa dòng {row_number} trong file local")
            return True
        except Exception as e:
            print(f"[DEBUG] Lỗi khi xóa dòng {row_number}: {e}")
            return False

    def update_transaction(self, row_number, transaction_type, category, amount, note='', purpose='', is_wife=False):
        """Cập nhật giao dịch theo số dòng trong sheet (không thay đổi cột Ngày)."""
        # Validate row number
        try:
            row_number = int(row_number)
        except (ValueError, TypeError):
            raise ValueError('ID giao dịch không hợp lệ')

        if row_number <= 1:
            raise ValueError('Không thể sửa header')

        # Validate dữ liệu giống add_transaction
        transaction_type = (transaction_type or '').strip()
        if transaction_type not in ['Thu', 'Chi']:
            raise ValueError('Loại giao dịch không hợp lệ')

        category = (category or '').strip()
        if not category:
            raise ValueError('Danh mục không được để trống')

        try:
            amount = float(amount)
            if amount <= 0:
                raise ValueError('Số tiền phải lớn hơn 0')
            if amount > 1e12:
                raise ValueError('Số tiền quá lớn')
        except (ValueError, TypeError):
            raise ValueError('Số tiền không hợp lệ')

        purpose = (purpose or '').strip()
        is_wife_flag = bool(is_wife)

        # Nếu là chi cho vợ mà bạn chưa ghi chú gì, tự điền 'Vợ'
        if is_wife_flag and not str(note).strip():
            note = 'Vợ'

        try:
            wb = self._load_local_workbook_for_write()
            if not wb or SHEET_NAME_INCOME_EXPENSE not in wb.sheetnames:
                raise ValueError('Không tìm thấy sheet ThuChi trong file local')
            ws = wb[SHEET_NAME_INCOME_EXPENSE]
            if row_number > ws.max_row:
                raise ValueError('ID giao dịch không tồn tại')

            # Đảm bảo có ít nhất 6 cột để không bị index error
            existing = list(ws.iter_rows(min_row=row_number, max_row=row_number, values_only=True))[0]
            row = list(existing) + [''] * (6 - len(existing))

            row[1] = transaction_type
            row[2] = category
            row[3] = f"{amount:.2f}" if amount < 1e10 else str(int(amount))
            row[4] = str(note).strip() if note else ''
            row[5] = purpose

            for idx, val in enumerate(row[:6], start=1):
                ws.cell(row=row_number, column=idx, value=val)
            self._save_local_workbook(wb)
            print(f"[DEBUG] Đã cập nhật dòng {row_number} trong file local")
            return True
        except Exception as e:
            print(f"[DEBUG] Lỗi khi cập nhật dòng {row_number}: {e}")
            raise

    def get_summary(self):
        """Lấy tổng kết thu chi"""
        transactions = self.get_all_transactions()
        
        total_income = 0
        total_expense = 0
        current_year = datetime.now().year
        counted_transactions = 0
        
        for t in transactions:
            try:
                # Chỉ tính giao dịch thuộc năm hiện tại
                date_str = t.get('date') or t.get('Ngày') or ''
                if date_str:
                    try:
                        tx_date = datetime.fromisoformat(str(date_str))
                    except ValueError:
                        # Thử parse nhanh theo format phổ biến dd/mm/yyyy hoặc dd-mm-yyyy
                        try:
                            tx_date = datetime.strptime(str(date_str), '%d/%m/%Y')
                        except ValueError:
                            try:
                                tx_date = datetime.strptime(str(date_str), '%d-%m-%Y')
                            except ValueError:
                                tx_date = None
                    if tx_date and tx_date.year != current_year:
                        continue

                amount = float(t.get('amount', 0))
                tx_type = t.get('type')
                category = t.get('category', '')
                is_wife = bool(t.get('is_wife'))

                if tx_type == 'Thu':
                    total_income += amount
                elif tx_type == 'Chi':
                    # Nếu là chi cho vợ và danh mục Quỹ thì bỏ qua khỏi tổng chi
                    if category == 'Quỹ' and is_wife:
                        continue
                    total_expense += amount
                counted_transactions += 1
            except (ValueError, TypeError):
                # Bỏ qua giao dịch có số tiền không hợp lệ
                continue
        
        balance = total_income - total_expense
        
        return {
            'total_income': total_income,
            'total_expense': total_expense,
            'balance': balance,
            'count': counted_transactions
        }

    def get_user_quy_summary(self, sheet_name: str):
        """
        Tính tổng số tiền quỹ cho từng loại quỹ (theo Mục đích) từ sheet cá nhân của user.
        - Lấy tất cả giao dịch type='Chi' và category='Quỹ'
        - Gom nhóm theo 'purpose' (Mục đích từ cột "Quỹ"). Nếu không có purpose thì gom vào 'Khác'.
        """
        if not sheet_name:
            return {'by_purpose': {}, 'total': 0.0}
        
        # Đọc tất cả records từ sheet của user
        records = self._get_records_from_local(sheet_name)
        if records is None:
            sheet = self.get_sheet(sheet_name)
            self._ensure_user_sheet_headers(sheet)
            records = sheet.get_all_records()
        
        quy_totals = {}
        total_all = 0.0

        for r in records:
            try:
                tx_type = (r.get('Loại') or '').strip()
                category = (r.get('Danh mục') or '').strip()
                if tx_type != 'Chi' or category.lower() != 'quỹ':
                    continue

                amount_raw = str(r.get('Số tiền') or '').replace('.', '').replace(',', '').strip()
                try:
                    amount = float(amount_raw)
                except (ValueError, TypeError):
                    continue
                if amount <= 0:
                    continue

                purpose = (r.get('Quỹ') or '').strip() or 'Khác'

                if purpose not in quy_totals:
                    quy_totals[purpose] = {'total': 0.0}

                quy_totals[purpose]['total'] += amount
                total_all += amount
            except (ValueError, TypeError):
                continue

        return {
            'by_purpose': quy_totals,
            'total': total_all
        }

    def get_combined_quy_summary(self, sheet_names: list):
        """
        Tính tổng số tiền quỹ từ nhiều sheet (ví dụ: Hưng và Giang).
        - Lấy tất cả giao dịch type='Chi' và category='Quỹ' từ các sheet
        - Gom nhóm theo 'purpose' (Mục đích từ cột "Quỹ"). Nếu không có purpose thì gom vào 'Khác'.
        - Trả về chi tiết theo từng người (sheet) và tổng.
        """
        quy_totals = {}
        total_all = 0.0
        totals_by_person = {}  # Tổng theo từng người

        for sheet_name in sheet_names:
            if not sheet_name:
                continue
            
            totals_by_person[sheet_name] = 0.0
            
            # Đọc tất cả records từ sheet
            records = self._get_records_from_local(sheet_name)
            if records is None:
                try:
                    sheet = self.get_sheet(sheet_name)
                    self._ensure_user_sheet_headers(sheet)
                    records = sheet.get_all_records()
                except Exception as e:
                    print(f"[WARN] Không đọc được sheet '{sheet_name}': {e}")
                    continue

            for r in records:
                try:
                    tx_type = (r.get('Loại') or '').strip()
                    category = (r.get('Danh mục') or '').strip()
                    if tx_type != 'Chi' or category.lower() != 'quỹ':
                        continue

                    amount_raw = str(r.get('Số tiền') or '').replace('.', '').replace(',', '').strip()
                    try:
                        amount = float(amount_raw)
                    except (ValueError, TypeError):
                        continue
                    if amount <= 0:
                        continue

                    purpose = (r.get('Quỹ') or '').strip() or 'Khác'

                    if purpose not in quy_totals:
                        quy_totals[purpose] = {'total': 0.0}
                        # Khởi tạo số tiền cho từng người
                        for name in sheet_names:
                            quy_totals[purpose][name] = 0.0

                    quy_totals[purpose][sheet_name] = quy_totals[purpose].get(sheet_name, 0.0) + amount
                    quy_totals[purpose]['total'] += amount
                    totals_by_person[sheet_name] += amount
                    total_all += amount
                except (ValueError, TypeError):
                    continue

        return {
            'by_purpose': quy_totals,
            'total': total_all,
            'totals_by_person': totals_by_person
        }

    def get_quy_summary(self):
        """
        Tính tổng số tiền quỹ cho từng loại quỹ (theo Mục đích).

        - Lấy tất cả giao dịch type='Chi' và category='Quỹ'
        - Gom nhóm theo 'purpose' (Mục đích). Nếu không có purpose thì gom vào 'Khác'.
        - Bao gồm cả giao dịch của vợ, vì đây là tiền quỹ độc lập với Tổng Chi.
        """
        transactions = self.get_all_transactions()

        quy_totals = {}
        total_all = 0.0
        total_husband = 0.0
        total_wife = 0.0

        for t in transactions:
            try:
                if t.get('type') != 'Chi':
                    continue
                if t.get('category') != 'Quỹ':
                    continue

                amount = float(t.get('amount', 0) or 0)
                if amount <= 0:
                    continue

                purpose = (t.get('purpose') or '').strip() or 'Khác'
                is_wife = bool(t.get('is_wife'))

                if purpose not in quy_totals:
                    quy_totals[purpose] = {
                        'total': 0.0,
                        'husband': 0.0,
                        'wife': 0.0
                    }

                if is_wife:
                    quy_totals[purpose]['wife'] += amount
                    total_wife += amount
                else:
                    quy_totals[purpose]['husband'] += amount
                    total_husband += amount

                quy_totals[purpose]['total'] += amount
                total_all += amount
            except (ValueError, TypeError):
                continue

        return {
            'by_purpose': quy_totals,
            'total': total_all,
            'total_husband': total_husband,
            'total_wife': total_wife
        }

    def get_categories(self):
        """Đọc danh mục từ file local (cột A: Thu, cột B: Chi) của sheet Danh Mục."""
        import re
        try:
            ws = self._get_local_ws(SHEET_NAME_CATEGORIES, read_only=True)
            if not ws:
                return {'Thu': [], 'Chi': []}
            self._log_source(SHEET_NAME_CATEGORIES, 'local')

            def read_col(idx):
                vals = []
                for row_idx, cell in enumerate(ws.iter_rows(min_col=idx, max_col=idx, values_only=True), start=1):
                    if row_idx == 1:
                        continue  # header
                    value = cell[0]
                    if value and str(value).strip():
                        vals.append(str(value).strip())
                return vals

            def parse(vals):
                items = []
                for v in vals:
                    match = re.match(r'^([\U0001F300-\U0001F9FF\U00002600-\U000027BF\U0001F1E0-\U0001F1FF\U0001F600-\U0001F64F\U0001F680-\U0001F6FF\U0001F900-\U0001F9FF\U00002700-\U000027BF\s]*)(.+)$', v)
                    if match:
                        icon = match.group(1).strip()
                        name = match.group(2).strip()
                    else:
                        icon = ''
                        name = v
                    if name:
                        items.append({'name': name, 'icon': icon})
                return items

            return {
                'Thu': parse(read_col(1)),
                'Chi': parse(read_col(2))
            }
        except Exception as e:
            print(f"[DEBUG] Lỗi get_categories: {e}")
            return {'Thu': [], 'Chi': []}

    def get_quy_purposes(self):
        """Lấy danh sách mục đích quỹ (cột C trong sheet Danh Mục), tách icon và tên từ cột C. Ưu tiên đọc từ local."""
        import re
        try:
            ws = self._get_local_ws(SHEET_NAME_CATEGORIES, read_only=True)
            if not ws:
                return []
            self._log_source(SHEET_NAME_CATEGORIES, 'local')
            purposes = []
            for idx, cell in enumerate(ws.iter_rows(min_col=3, max_col=3, values_only=True), start=1):
                if idx == 1:
                    continue  # header
                value = cell[0]
                if value and str(value).strip():
                    full_value = str(value).strip()
                    match = re.match(r'^([\U0001F300-\U0001F9FF\U00002600-\U000027BF\U0001F1E0-\U0001F1FF\U0001F600-\U0001F64F\U0001F680-\U0001F6FF\U0001F900-\U0001F9FF\U00002700-\U000027BF\s]*)(.+)$', full_value)
                    if match:
                        icon = match.group(1).strip()
                        purpose_name = match.group(2).strip()
                    else:
                        icon = ''
                        purpose_name = full_value
                    if purpose_name:
                        purposes.append({'name': purpose_name, 'icon': icon})
            return purposes
        except Exception as e:
            print(f"[DEBUG] Lỗi khi lấy mục đích quỹ: {e}")
            import traceback
            traceback.print_exc()
            return []

    def get_available_icons(self):
        """Lấy danh sách icon có sẵn từ cột D và E trong sheet DanhMục (bỏ header, bỏ trùng)."""
        try:
            ws = self._get_local_ws(SHEET_NAME_CATEGORIES, read_only=True)
            if not ws:
                return []
            icons = []
            seen = set()
            for col_idx in (4, 5):  # D, E
                for idx, cell in enumerate(ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True), start=1):
                    if idx == 1:
                        continue  # header
                    value = cell[0]
                    if value and str(value).strip():
                        icon = str(value).strip()
                        if icon and icon not in seen:
                            icons.append(icon)
                            seen.add(icon)
            return icons
        except Exception as e:
            print(f"[DEBUG] Lỗi khi lấy icon: {e}")
            return []

    # ----------------- CRUD danh mục -----------------
    def update_category(self, category_type: str, old_name: str, new_name: str, icon: str = ''):
        """Cập nhật danh mục theo tên cũ -> tên mới (cùng loại)."""
        ws = self._get_local_ws(SHEET_NAME_CATEGORIES, read_only=False, create=True)
        if not ws:
            raise ValueError("Không mở được sheet Danh Mục")
        col_map = {'Thu': 1, 'Chi': 2, 'Quỹ': 3}
        col = col_map.get(category_type)
        if not col:
            raise ValueError("Loại danh mục không hợp lệ")

        target_row = None
        for idx, cell in enumerate(ws.iter_rows(min_col=col, max_col=col), start=1):
            if idx == 1:
                continue
            val = cell[0].value if cell and cell[0] else None
            if val and str(val).strip().endswith(old_name):
                target_row = idx
                break
        if not target_row:
            return False
        new_val = f"{icon} {new_name}".strip()
        ws.cell(row=target_row, column=col, value=new_val)
        self._save_local_workbook(ws.parent)
        return True

    def delete_category(self, category_type: str, name: str):
        """Xóa danh mục theo tên (cùng loại)."""
        ws = self._get_local_ws(SHEET_NAME_CATEGORIES, read_only=False, create=True)
        if not ws:
            raise ValueError("Không mở được sheet Danh Mục")
        col_map = {'Thu': 1, 'Chi': 2, 'Quỹ': 3}
        col = col_map.get(category_type)
        if not col:
            raise ValueError("Loại danh mục không hợp lệ")

        target_row = None
        for idx, cell in enumerate(ws.iter_rows(min_col=col, max_col=col), start=1):
            if idx == 1:
                continue
            val = cell[0].value if cell and cell[0] else None
            if val and str(val).strip().endswith(name):
                target_row = idx
                break
        if not target_row:
            return False
        ws.delete_rows(target_row, 1)
        self._save_local_workbook(ws.parent)
        return True

    def add_category(self, category_type, category_name, icon=''):
        """Thêm danh mục mới vào sheet DanhMuc
        
        Args:
            category_type: 'Thu', 'Chi', hoặc 'Quỹ'
            category_name: Tên danh mục cần thêm
            icon: Icon cho danh mục (tùy chọn, sẽ kết hợp với tên danh mục)
        
        Returns:
            True nếu thành công, False nếu danh mục đã tồn tại
        """
        if category_type not in ['Thu', 'Chi', 'Quỹ']:
            raise ValueError('Loại danh mục không hợp lệ')
        
        category_name = (category_name or '').strip()
        if not category_name:
            raise ValueError('Tên danh mục không được để trống')
        
        icon = (icon or '').strip()

        # Xác định cột dựa trên loại
        if category_type == 'Quỹ':
            column_letter = 'C'  # Cột C cho mục đích quỹ
        else:
            column_letter = CATEGORY_COLUMNS.get(category_type)
            if not column_letter:
                raise ValueError(f'Không tìm thấy cột cho loại {category_type}')

        ws = self._get_local_ws(SHEET_NAME_CATEGORIES, read_only=False, create=True)
        wb = ws.parent
        # Nếu sheet mới và trống, thêm header mặc định
        if ws.max_row == 0:
            ws.append(['Danh mục Thu', 'Danh mục Chi', 'Mục đích Quỹ', 'Icon D', 'Icon E'])

        _, col_index = a1_to_rowcol(f"{column_letter}1")
        col_values = [cell[0] for cell in ws.iter_rows(min_col=col_index, max_col=col_index, values_only=True)]

        # Tách icon và tên từ các danh mục hiện có để so sánh
        existing_categories = []
        import re
        for value in col_values[1:]:
            if value and str(value).strip():
                cat_name = str(value).strip()
                cat_name_clean = re.sub(
                    r'^[\U0001F300-\U0001F9FF\U00002600-\U000027BF\U0001F1E0-\U0001F1FF\U0001F600-\U0001F64F\U0001F680-\U0001F6FF\U0001F900-\U0001F9FF\U00002700-\U000027BF\s]+',
                    '',
                    cat_name
                )
                if not cat_name_clean:
                    cat_name_clean = cat_name
                existing_categories.append(cat_name_clean.strip().lower())

        # Kiểm tra xem danh mục đã tồn tại chưa (không phân biệt hoa thường, bỏ qua icon)
        if category_name.lower() in existing_categories:
            return False

        # Tìm dòng cuối cùng có dữ liệu trong cột này
        last_row = 1  # Bắt đầu từ header
        for idx, value in enumerate(col_values[1:], start=2):  # Bỏ qua header (index 0)
            if value and str(value).strip():
                last_row = idx
        
        # Thêm vào dòng tiếp theo sau dòng cuối cùng có dữ liệu
        next_row = last_row + 1

        category_value = f"{icon} {category_name}" if icon else category_name
        ws.cell(row=next_row, column=col_index, value=category_value)
        self._save_local_workbook(wb)
        print(f"[DEBUG] Đã thêm danh mục '{category_value}' vào file local, cột {column_letter}, dòng {next_row}")

        return True


# Singleton instance
_sheets_manager = None

def get_sheets_manager():
    """Lấy instance của SheetsManager (singleton)"""
    global _sheets_manager
    if _sheets_manager is None:
        _sheets_manager = SheetsManager()
    return _sheets_manager

